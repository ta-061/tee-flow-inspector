#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OP-TEE TA 解析結果メトリクス収集ツール（日本語Excel出力・列幅自動調整・DITING厳密/ゆるやか一致・time.txt対応）

機能概要:
- benchmark/<プロジェクト名>/ta/results 配下の成果物を横断収集
- ta_vulnerabilities.json / ta_candidate_flows.json / ta_chains.json / ta_call_graph.json /
  ta_sinks.json / ta_vulnerable_destinations.json / taint_analysis_log.txt / time.txt 等から
  脆弱性件数・チェイン数・候補フロー数・トークン/時間・キャッシュ統計・タイムスタンプなどを抽出
- DITING_ans.csv（任意）を読み込み、(file,line[,sink]) の一致を計算
  * sink列が無い場合は (file,line) で一致判定
  * 行番号は ±N 行の許容範囲でゆるやか一致（--diting-line-tol, 既定=3）
  * DITINGが start_line/end_line を持つ場合は start_line を採用
- 日本語カラム・日本語シート名で Excel を出力（列幅は各列の最長表示幅に自動調整）
- 出力: 既定で src/metrics/analysis_metrics.xlsx

使い方:
  python3 src/metrics/collect_metrics.py \
    --benchmark-root benchmark \
    --out src/metrics/analysis_metrics.xlsx \
    --diting src/metrics/DITING_ans.csv \
    --diting-line-tol 5
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Set, DefaultDict
from collections import defaultdict

import pandas as pd

# ------------------------------------------------------------
# ユーティリティ
# ------------------------------------------------------------

def safe_read_json(path: Path) -> Optional[Any]:
    try:
        if path.is_file():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return None

def safe_read_text(path: Path) -> Optional[str]:
    try:
        if path.is_file():
            return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None
    return None

def text_display_len(s: Any) -> int:
    """
    Excel列幅調整用の見かけ文字幅。CJKは2、その他は1で概算。
    """
    if s is None:
        return 0
    t = str(s)
    width = 0
    for ch in t:
        o = ord(ch)
        if (
            0x2E80 <= o <= 0x9FFF  # 中日韓
            or 0xAC00 <= o <= 0xD7AF  # 韓
            or 0x3040 <= o <= 0x30FF  # ひら・カタカナ
            or 0xFF01 <= o <= 0xFF60  # 全角記号
            or 0xFFE0 <= o <= 0xFFE6
        ):
            width += 2
        else:
            width += 1
    return width

def autosize_all_columns(ws):
    """
    openpyxl ワークシートの全列幅をセルの最長表示幅に合わせて調整。
    """
    from openpyxl.utils import get_column_letter

    max_widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_idx = cell.column
            w = text_display_len(cell.value)
            if w > max_widths.get(col_idx, 0):
                max_widths[col_idx] = w

    for col_idx, w in max_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(8, min(120, w + 2))

def write_df(ws_name: str, df: pd.DataFrame, writer: pd.ExcelWriter, freeze: str = "A2", apply_autofilter: bool = True):
    df.to_excel(writer, sheet_name=ws_name, index=False)
    ws = writer.book[ws_name]
    if apply_autofilter:
        ws.auto_filter.ref = ws.dimensions
    if freeze:
        ws.freeze_panes = freeze
    autosize_all_columns(ws)

def ensure_dir(p: Path):
    p.parent.mkdir(parents=True, exist_ok=True)

def detect_project_and_prefix(results_dir: Path) -> Tuple[str, str]:
    """
    期待パス: benchmark/<project>/ta/results
    - プロジェクト名: 'ta' ディレクトリの一つ上
    - 成果物プレフィックス: 直上ディレクトリ名（通常 'ta'）
    """
    ta_dir = results_dir.parent        # .../<project>/ta
    proj_dir = ta_dir.parent           # .../<project>
    project = proj_dir.name
    prefix = ta_dir.name               # 通常 'ta'
    return project, prefix

# ------------------------------------------------------------
# 解析結果パーサ
# ------------------------------------------------------------

@dataclass
class ProjectMetrics:
    # 表示/識別用
    project_name: str
    results_dir: Path
    artifact_prefix: str  # 多くの成果物ファイル名の接頭辞（通常 'ta'）

    # 件数系
    vuln_count: int = 0
    candidate_flows: int = 0
    chains: int = 0
    vd_calls: int = 0
    sinks_count: int = 0
    callgraph_edges: int = 0
    defined_functions: int = 0

    # トークン/時間（ログや JSON）
    api_calls: Optional[int] = None
    prompt_tokens: Optional[int] = None
    completion_tokens: Optional[int] = None
    total_tokens: Optional[int] = None
    analysis_seconds: Optional[float] = None
    analysis_time_label: Optional[str] = None
    cache_hits: Optional[int] = None
    cache_misses: Optional[int] = None
    cache_hit_rate: Optional[float] = None

    # time.txt 由来（存在すれば、analysis_seconds が未設定のとき補完）
    time_start: Optional[str] = None
    time_end: Optional[str] = None
    time_seconds: Optional[float] = None
    time_label: Optional[str] = None
    analysis_mode: Optional[str] = None
    token_tracking: Optional[str] = None  # "Enabled"/"Disabled"
    phase_rows: List[Dict[str, Any]] = field(default_factory=list)

    # 分布など
    severity_hist: Dict[str, int] = field(default_factory=dict)
    cwe_hist: Dict[str, int] = field(default_factory=dict)

    # DITING比較
    diting_count: Optional[int] = None
    match_count: Optional[int] = None
    match_rate: Optional[float] = None

    # リンク
    report_html: Optional[Path] = None

    # 検出詳細
    vuln_rows: List[Dict[str, Any]] = field(default_factory=list)

# ---- ta_vulnerabilities.json ----

def parse_vuln_json(vuln_json: Any, pm: ProjectMetrics):
    if vuln_json is None:
        return

    # 統計
    stats = vuln_json.get("statistics") if isinstance(vuln_json, dict) else None
    if isinstance(stats, dict):
        tok = stats.get("token_usage", {})
        pm.api_calls = tok.get("api_calls", pm.api_calls)
        pm.prompt_tokens = tok.get("total_prompt_tokens", pm.prompt_tokens)
        pm.completion_tokens = tok.get("total_completion_tokens", pm.completion_tokens)
        pm.total_tokens = tok.get("total_tokens", pm.total_tokens)

        pm.analysis_seconds = stats.get("analysis_time_seconds", pm.analysis_seconds)
        pm.analysis_time_label = stats.get("analysis_time_formatted", pm.analysis_time_label)

        # キャッシュ統計（新形式）
        cst = stats.get("cache_stats", {})
        if isinstance(cst, dict):
            pm.cache_hits = cst.get("hits", pm.cache_hits)
            pm.cache_misses = cst.get("misses", pm.cache_misses)
            hit_rate = cst.get("hit_rate")
            if isinstance(hit_rate, str) and hit_rate.endswith("%"):
                try:
                    pm.cache_hit_rate = float(hit_rate[:-1])
                except Exception:
                    pass
            elif isinstance(hit_rate, (int, float)):
                pm.cache_hit_rate = float(hit_rate)

    # 脆弱性配列
    vulns = None
    if isinstance(vuln_json, dict) and "vulnerabilities" in vuln_json:
        vulns = vuln_json.get("vulnerabilities", [])
    elif isinstance(vuln_json, list):
        vulns = vuln_json

    if isinstance(vulns, list):
        pm.vuln_count = len(vulns)
        for v in vulns:
            vd = v.get("vd", {})
            file_path = vd.get("file") or v.get("file")
            line = vd.get("line") or v.get("line")
            sink = vd.get("sink") or v.get("sink")

            details = v.get("vulnerability_details", {})
            det = details.get("details", details) if isinstance(details, dict) else {}
            cwe = det.get("vulnerability_type") or v.get("vulnerability_type")
            sev = det.get("severity") or v.get("severity")
            flow_summary = det.get("taint_flow_summary", {})
            chain = None
            if isinstance(flow_summary, dict) and "propagation_path" in flow_summary:
                chain = " -> ".join(flow_summary["propagation_path"])

            sev_key = sev or "unknown"
            cwe_key = cwe or "unknown"
            pm.severity_hist[sev_key] = pm.severity_hist.get(sev_key, 0) + 1
            pm.cwe_hist[cwe_key] = pm.cwe_hist.get(cwe_key, 0) + 1

            pm.vuln_rows.append({
                "プロジェクト": pm.project_name,
                "ファイル": file_path,
                "行": line,
                "シンク": sink,
                "CWE": cwe,
                "深刻度": sev,
                "チェイン": chain,
            })

# ---- taint_analysis_log.txt ----

def parse_log_for_tokens_and_cache(log_text: str, pm: ProjectMetrics):
    if not log_text:
        return

    # 所要時間: "所要時間: 6.4分"
    m = re.search(r"所要時間:\s*([0-9.]+)\s*分", log_text)
    if m and pm.analysis_seconds is None:
        minutes = float(m.group(1))
        pm.analysis_seconds = minutes * 60.0
        pm.analysis_time_label = f"{minutes:.1f}分"

    # LLM呼び出し回数
    m = re.search(r"LLM呼び出し回数:\s*([0-9,]+)", log_text)
    if m and pm.api_calls is None:
        pm.api_calls = int(m.group(1).replace(",", ""))

    # 総トークン数
    m = re.search(r"総トークン数:\s*([0-9,]+)", log_text)
    if m and pm.total_tokens is None:
        pm.total_tokens = int(m.group(1).replace(",", ""))

    # 入力/出力トークン
    m = re.search(r"入力トークン:\s*([0-9,]+)", log_text)
    if m and pm.prompt_tokens is None:
        pm.prompt_tokens = int(m.group(1).replace(",", ""))
    m = re.search(r"出力トークン:\s*([0-9,]+)", log_text)
    if m and pm.completion_tokens is None:
        pm.completion_tokens = int(m.group(1).replace(",", ""))

    # キャッシュ統計
    mh = re.search(r"ヒット数:\s*([0-9,]+)", log_text)
    mm = re.search(r"ミス数:\s*([0-9,]+)", log_text)
    mr = re.search(r"ヒット率:\s*([0-9.]+)%", log_text)
    if mh and pm.cache_hits is None:
        pm.cache_hits = int(mh.group(1).replace(",", ""))
    if mm and pm.cache_misses is None:
        pm.cache_misses = int(mm.group(1).replace(",", ""))
    if mr and pm.cache_hit_rate is None:
        pm.cache_hit_rate = float(mr.group(1))

# ---- results_dir/time.txt ----

_TIME_KV_RE = re.compile(r"^\s*([A-Za-z ]+):\s*(.+?)\s*$")
_PHASE_RE = re.compile(r"^\s*(.+?):\s*([0-9.]+)s\s*\(\s*([0-9.]+)%\s*\)\s*$")
_DURATION_RE = re.compile(r"^\s*(\d+)m\s*([0-9.]+)s\s*$")

def _parse_duration_to_seconds(label: str) -> Optional[float]:
    m = _DURATION_RE.match(label.strip())
    if not m:
        return None
    mins = int(m.group(1))
    secs = float(m.group(2))
    return mins * 60.0 + secs

def parse_time_txt(text: str, pm: ProjectMetrics):
    """
    例: time.txt
    Analysis Mode: Hybrid (DITING rules only)
    Token Tracking: Enabled
    Start Time: 2025-08-15 06:52:36
    End Time:   2025-08-15 06:56:36
    Total Duration: 3m 59.56s
    Total Seconds: 239.56s
    Phase Breakdown:
      phase6_taint_analysis             : 2m 9.31s        ( 54.0%)
    """
    if not text:
        return

    lines = [ln.rstrip("\n") for ln in text.splitlines()]
    in_phase = False
    for ln in lines:
        if ln.strip().lower().startswith("phase breakdown"):
            in_phase = True
            continue

        if not in_phase:
            m = _TIME_KV_RE.match(ln)
            if not m:
                continue
            key = m.group(1).strip().lower()
            val = m.group(2).strip()

            if key.startswith("analysis mode"):
                pm.analysis_mode = val
            elif key.startswith("token tracking"):
                pm.token_tracking = val
            elif key.startswith("start time"):
                pm.time_start = val
            elif key.startswith("end time"):
                pm.time_end = val
            elif key.startswith("total seconds"):
                # "239.56s" -> 239.56
                v = val.rstrip("s")
                try:
                    pm.time_seconds = float(v)
                except Exception:
                    pass
            elif key.startswith("total duration"):
                pm.time_label = val
                # 秒が空で、ラベルが "Xm Y.s" 形式なら秒も埋める
                if pm.time_seconds is None:
                    sec = _parse_duration_to_seconds(val)
                    if sec is not None:
                        pm.time_seconds = sec
        else:
            # フェーズ行
            m = _PHASE_RE.match(ln.replace(" ", ""))  # スペース詰めてから一致させる
            if not m:
                # より寛容に: "phase6_taint_analysis : 2m 9.31s ( 54.0%)" のような表記
                ln2 = ln.strip()
                # "name : 2m 9.31s (54.0%)" を取り出す
                m2 = re.match(r"^(.+?):\s*([0-9m.\s]+s)\s*\(\s*([0-9.]+)%\s*\)\s*$", ln2)
                if m2:
                    name = m2.group(1).strip()
                    label = m2.group(2).strip()
                    ratio = float(m2.group(3))
                    seconds = _parse_duration_to_seconds(label) if "m" in label else float(label.rstrip("s"))
                    pm.phase_rows.append({"プロジェクト": pm.project_name, "フェーズ": name, "秒(time.txt)": seconds, "割合%(time.txt)": ratio})
                continue
            # m が取れたときは name は取れないのでスキップ（寛容モードに委譲）

    # time.txt の秒があり、analysis_seconds が未設定なら補完
    if pm.time_seconds is not None and pm.analysis_seconds is None:
        pm.analysis_seconds = pm.time_seconds
        if pm.time_label and pm.analysis_time_label is None:
            pm.analysis_time_label = pm.time_label

# ------------------------------------------------------------
# 各プロジェクトの収集
# ------------------------------------------------------------

def collect_for_project(results_dir: Path) -> ProjectMetrics:
    project, prefix = detect_project_and_prefix(results_dir)
    pm = ProjectMetrics(project_name=project, results_dir=results_dir, artifact_prefix=prefix)

    # 成果物パス
    f_vuln = results_dir / f"{prefix}_vulnerabilities.json"
    f_cands = results_dir / f"{prefix}_candidate_flows.json"
    f_chains = results_dir / f"{prefix}_chains.json"
    f_callg = results_dir / f"{prefix}_call_graph.json"
    f_sinks = results_dir / f"{prefix}_sinks.json"
    f_vds   = results_dir / f"{prefix}_vulnerable_destinations.json"
    f_log   = results_dir / "taint_analysis_log.txt"
    f_html  = results_dir / f"{prefix}_vulnerability_report.html"
    f_time  = results_dir / "time.txt"

    # レポートリンク
    if f_html.is_file():
        pm.report_html = f_html

    # JSON系
    j_vuln = safe_read_json(f_vuln)
    j_cands = safe_read_json(f_cands)
    j_chains = safe_read_json(f_chains)
    j_callg = safe_read_json(f_callg)
    j_sinks = safe_read_json(f_sinks)
    j_vds   = safe_read_json(f_vds)

    # 件数集計
    if isinstance(j_cands, list):
        pm.candidate_flows = len(j_cands)
    if isinstance(j_chains, list):
        pm.chains = len(j_chains)
    if isinstance(j_vds, list):
        pm.vd_calls = len(j_vds)
    if isinstance(j_sinks, dict) and "sinks" in j_sinks:
        pm.sinks_count = len(j_sinks["sinks"])
    elif isinstance(j_sinks, list):
        pm.sinks_count = len(j_sinks)
    if isinstance(j_callg, dict) and "edges" in j_callg:
        pm.callgraph_edges = len(j_callg["edges"])
        if "definitions" in j_callg and isinstance(j_callg["definitions"], dict):
            pm.defined_functions = len(j_callg["definitions"])

    # 脆弱性本体/統計
    parse_vuln_json(j_vuln, pm)

    # ログから補完
    if f_log.is_file():
        parse_log_for_tokens_and_cache(safe_read_text(f_log) or "", pm)

    # time.txt
    if f_time.is_file():
        parse_time_txt(safe_read_text(f_time) or "", pm)

    return pm

def scan_results(benchmark_root: Path) -> List[ProjectMetrics]:
    metrics: List[ProjectMetrics] = []
    for ta_results in sorted(benchmark_root.glob("*/ta/results")):
        if not ta_results.is_dir():
            continue
        metrics.append(collect_for_project(ta_results))
    return metrics

# ------------------------------------------------------------
# DITING 比較
# ------------------------------------------------------------

def normalize_proj_name(s: str) -> str:
    return Path(str(s)).name.lower().strip()

def normalize_file_basename(s: str) -> str:
    return Path(str(s)).name.lower().strip()

def load_diting_csv(diting_path: Path) -> Optional[pd.DataFrame]:
    if not diting_path.is_file():
        return None
    try:
        df = pd.read_csv(diting_path)
        # 列名を標準化（小文字化）
        df.columns = [c.strip().lower() for c in df.columns]
        return df
    except Exception:
        return None

def pick_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None

def extract_vd_triplets(pm: ProjectMetrics) -> List[Tuple[str,int,Optional[str]]]:
    """
    我々側の (file,line,sink?) 三つ組を results から抽出。
    優先: *_vulnerable_destinations.json, 予備: *_chains.json, さらに ta_vulnerabilities.json の vd / inline_findings も採用。
    """
    triplets: List[Tuple[str,int,Optional[str]]] = []
    prefix = pm.artifact_prefix

    # 1) vulnerable_destinations.json
    f_vds = pm.results_dir / f"{prefix}_vulnerable_destinations.json"
    if f_vds.is_file():
        try:
            vds = json.loads(f_vds.read_text(encoding="utf-8"))
            if isinstance(vds, list):
                for e in vds:
                    vd = e.get("vd", e) if isinstance(e, dict) else {}
                    file_ = vd.get("file"); line_ = vd.get("line"); sink_ = vd.get("sink")
                    if file_ and line_ is not None:
                        triplets.append((normalize_file_basename(file_), int(line_), str(sink_) if sink_ else None))
        except Exception:
            pass

    # 2) chains.json（予備）
    if not triplets:
        f_chains = pm.results_dir / f"{prefix}_chains.json"
        if f_chains.is_file():
            try:
                chains = json.loads(f_chains.read_text(encoding="utf-8"))
                if isinstance(chains, list):
                    for e in chains:
                        vd = e.get("vd", {}) if isinstance(e, dict) else {}
                        file_ = vd.get("file"); line_ = vd.get("line"); sink_ = vd.get("sink")
                        if file_ and line_ is not None:
                            triplets.append((normalize_file_basename(file_), int(line_), str(sink_) if sink_ else None))
            except Exception:
                pass

    # 3) ta_vulnerabilities.json（補完: vd と inline_findings の行を追加）
    f_vuln = pm.results_dir / f"{prefix}_vulnerabilities.json"
    if f_vuln.is_file():
        try:
            vj = json.loads(f_vuln.read_text(encoding="utf-8"))
            # a) vulnerabilities[].vd
            vulns = vj.get("vulnerabilities", []) if isinstance(vj, dict) else (vj if isinstance(vj, list) else [])
            if isinstance(vulns, list):
                for v in vulns:
                    vd = v.get("vd", {}) if isinstance(v, dict) else {}
                    file_ = vd.get("file"); line_ = vd.get("line"); sink_ = vd.get("sink")
                    if file_ and line_ is not None:
                        triplets.append((normalize_file_basename(file_), int(line_), str(sink_) if sink_ else None))
            # b) inline_findings[]
            inls = vj.get("inline_findings", []) if isinstance(vj, dict) else []
            if isinstance(inls, list):
                for it in inls:
                    file_ = it.get("file"); line_ = it.get("line"); sink_ = it.get("sink_function") or it.get("sink")
                    if file_ and line_ is not None:
                        triplets.append((normalize_file_basename(file_), int(line_), str(sink_) if sink_ else None))
        except Exception:
            pass

    # 重複除去
    triplets = list(dict.fromkeys(triplets))
    return triplets

def compute_matches(pm: ProjectMetrics, diting_df: pd.DataFrame, vd_list: List[Tuple[str,int,Optional[str]]], line_tol: int = 3):
    """
    DITING の (project,file,start_line[,sink]) と我々の (file,line[,sink]) を突き合わせ。
    - sink列が無い場合は (file,line) のみで判定
    - 行は ±line_tol の許容
    """
    if diting_df is None or diting_df.empty:
        return

    col_proj = pick_column(diting_df, ["project", "ta", "name"])
    col_file = pick_column(diting_df, ["file", "source", "path"])
    col_line = pick_column(diting_df, ["line", "lineno", "start_line"])
    col_sink = pick_column(diting_df, ["sink", "function", "callee", "sink_function"])
    col_end  = pick_column(diting_df, ["end_line"])

    if col_proj is None or col_file is None or col_line is None:
        return

    # 対象プロジェクト（プロジェクト名は 'ta' の親ディレクトリ）
    dsub = diting_df[diting_df[col_proj].apply(lambda x: normalize_proj_name(str(x)) == pm.project_name.lower())].copy()
    pm.diting_count = int(dsub.shape[0])

    if pm.diting_count == 0:
        pm.match_count = 0
        pm.match_rate = 0.0
        return

    # 我々側の索引: file -> [(line, sink)]
    file2rows: DefaultDict[str, List[Tuple[int, Optional[str]]]] = defaultdict(list)
    for f, l, s in vd_list:
        file2rows[normalize_file_basename(f)].append((int(l), s if s else None))

    matches = 0
    for _, r in dsub.iterrows():
        f = normalize_file_basename(r[col_file])
        # line の決定: start_line を優先（無ければ line/lineno）
        try:
            lstart = int(r[col_line])
        except Exception:
            continue
        # sink（任意）
        s = str(r[col_sink]) if (col_sink and pd.notna(r[col_sink])) else None

        rows = file2rows.get(f, [])
        if not rows:
            continue

        if s:
            # sink も一致させる
            hit = any((abs(l - lstart) <= line_tol and (sr is None or str(sr) == s)) for (l, sr) in rows)
        else:
            # sink 不明なら (file,line) 近傍マッチ
            hit = any(abs(l - lstart) <= line_tol for (l, _sr) in rows)
        if hit:
            matches += 1

    pm.match_count = matches
    pm.match_rate = (matches / pm.diting_count * 100.0) if pm.diting_count else 0.0

# ------------------------------------------------------------
# メイン集計 → 日本語Excel
# ------------------------------------------------------------

def build_overview_df(projects: List[ProjectMetrics]) -> pd.DataFrame:
    n = len(projects)
    total_vuln = sum(p.vuln_count for p in projects)
    total_cands = sum(p.candidate_flows for p in projects)
    total_chains = sum(p.chains for p in projects)
    total_vds = sum(p.vd_calls for p in projects)
    total_edges = sum(p.callgraph_edges for p in projects)
    total_funcs = sum(p.defined_functions for p in projects)

    # トークン/時間は存在するもののみ合計・平均
    api_calls_list = [p.api_calls for p in projects if p.api_calls is not None]
    total_toks  = [p.total_tokens for p in projects if p.total_tokens is not None]
    secs        = [p.analysis_seconds for p in projects if p.analysis_seconds is not None]

    total_api_calls = sum(api_calls_list) if api_calls_list else None
    sum_total_toks  = sum(total_toks) if total_toks else None
    sum_secs        = sum(secs) if secs else None

    avg_tokens_per_call = None
    if sum_total_toks is not None and total_api_calls:
        avg_tokens_per_call = sum_total_toks / total_api_calls if total_api_calls else None

    avg_secs = (sum_secs / len(secs)) if secs else None

    # DITING
    d_total = sum(p.diting_count or 0 for p in projects)
    d_match = sum(p.match_count or 0 for p in projects)
    d_rate  = (d_match / d_total * 100.0) if d_total else 0.0

    rows = [
        {"項目": "解析対象プロジェクト数", "値": n},
        {"項目": "総脆弱性数(LLM)", "値": total_vuln},
        {"項目": "総候補フロー数", "値": total_cands},
        {"項目": "総チェイン数", "値": total_chains},
        {"項目": "総シンク呼び出し箇所(VD)", "値": total_vds},
        {"項目": "総関数数", "値": total_funcs},
        {"項目": "総呼び出しエッジ数", "値": total_edges},
        {"項目": "総API呼び出し回数(LLM)", "値": total_api_calls},
        {"項目": "総トークン数", "値": sum_total_toks},
        {"項目": "平均トークン数/呼び出し", "値": f"{avg_tokens_per_call:.2f}" if avg_tokens_per_call else None},
        {"項目": "総解析時間(秒)", "値": int(sum_secs) if sum_secs else None},
        {"項目": "平均解析時間(秒)", "値": f"{avg_secs:.1f}" if avg_secs else None},
        {"項目": "DITING総件数", "値": d_total},
        {"項目": "一致件数合計", "値": d_match},
        {"項目": "一致率(%)", "値": f"{d_rate:.1f}"},
    ]
    return pd.DataFrame(rows)

def build_per_project_df(projects: List[ProjectMetrics]) -> pd.DataFrame:
    rows = []
    for p in projects:
        rows.append({
            "プロジェクト": p.project_name,
            "脆弱性数(LLM)": p.vuln_count,
            "候補フロー数": p.candidate_flows,
            "チェイン数": p.chains,
            "シンク呼び出し箇所数(VD)": p.vd_calls,
            "シンク関数数": p.sinks_count,
            "関数数": p.defined_functions,
            "呼び出しエッジ数": p.callgraph_edges,
            "API呼び出し回数(LLM)": p.api_calls,
            "総トークン数": p.total_tokens,
            "解析時間(秒)": int(p.analysis_seconds) if p.analysis_seconds else None,
            "解析時間(表記)": p.analysis_time_label,
            "キャッシュ:ヒット数": p.cache_hits,
            "キャッシュ:ミス数": p.cache_misses,
            "キャッシュ:ヒット率(%)": p.cache_hit_rate,
            "DITING件数": p.diting_count,
            "DITING一致件数": p.match_count,
            "DITING一致率(%)": f"{p.match_rate:.1f}" if p.match_rate is not None else None,
            "レポートHTML": str(p.report_html) if p.report_html else None,
            # time.txt 由来
            "開始時刻(time.txt)": p.time_start,
            "終了時刻(time.txt)": p.time_end,
            "解析時間(秒, time.txt)": int(p.time_seconds) if p.time_seconds else None,
            "解析時間(表記, time.txt)": p.time_label,
            "分析モード(time.txt)": p.analysis_mode,
            "TokenTracking(time.txt)": p.token_tracking,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(by=["プロジェクト"]).reset_index(drop=True)
    return df

def build_token_time_df(projects: List[ProjectMetrics]) -> pd.DataFrame:
    rows = []
    for p in projects:
        rows.append({
            "プロジェクト": p.project_name,
            "API呼び出し回数(LLM)": p.api_calls,
            "総トークン数": p.total_tokens,
            "平均トークン数/呼び出し": (p.total_tokens / p.api_calls) if (p.total_tokens and p.api_calls) else None,
            "解析時間(秒)": int(p.analysis_seconds) if p.analysis_seconds else None,
            "解析時間(表記)": p.analysis_time_label,
            "キャッシュ:ヒット数": p.cache_hits,
            "キャッシュ:ミス数": p.cache_misses,
            "キャッシュ:ヒット率(%)": p.cache_hit_rate,
            # time.txt 由来
            "開始時刻(time.txt)": p.time_start,
            "終了時刻(time.txt)": p.time_end,
            "解析時間(秒, time.txt)": int(p.time_seconds) if p.time_seconds else None,
            "解析時間(表記, time.txt)": p.time_label,
            "分析モード(time.txt)": p.analysis_mode,
            "TokenTracking(time.txt)": p.token_tracking,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(by=["プロジェクト"]).reset_index(drop=True)
    return df

def build_diting_compare_df(projects: List[ProjectMetrics]) -> pd.DataFrame:
    rows = []
    for p in projects:
        rows.append({
            "プロジェクト": p.project_name,
            "DITING件数": p.diting_count,
            "一致件数": p.match_count,
            "一致率(%)": f"{p.match_rate:.1f}" if p.match_rate is not None else None,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(by=["プロジェクト"]).reset_index(drop=True)
    return df

def build_vuln_detail_df(projects: List[ProjectMetrics]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for p in projects:
        rows.extend(p.vuln_rows)
    df = pd.DataFrame(rows, columns=["プロジェクト","ファイル","行","シンク","CWE","深刻度","チェイン"])
    if not df.empty:
        if "行" in df.columns:
            df["行"] = pd.to_numeric(df["行"], errors="coerce").astype("Int64")
        df = df.sort_values(by=["プロジェクト","ファイル","行","シンク"], na_position="last").reset_index(drop=True)
    return df

def build_phase_breakdown_df(projects: List[ProjectMetrics]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for p in projects:
        rows.extend(p.phase_rows)
    df = pd.DataFrame(rows, columns=["プロジェクト","フェーズ","秒(time.txt)","割合%(time.txt)"])
    if not df.empty:
        df = df.sort_values(by=["プロジェクト","フェーズ"]).reset_index(drop=True)
    return df

# ------------------------------------------------------------
# CLI
# ------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="TA解析メトリクスの日本語Excel集計（DITING/line許容・time.txt対応）")
    ap.add_argument("--benchmark-root", type=Path, default=Path("benchmark"), help="benchmark ルートディレクトリ")
    ap.add_argument("--out", type=Path, default=Path("src/metrics/analysis_metrics.xlsx"), help="出力Excelパス")
    ap.add_argument("--diting", type=Path, default=Path("src/metrics/DITING_ans.csv"), help="DITING比較CSV（任意）")
    ap.add_argument("--diting-line-tol", type=int, default=3, help="DITING一致の行番号許容（±N行、既定=3）")
    args = ap.parse_args()

    # 1) プロジェクト横断収集
    projects = scan_results(args.benchmark_root)

    # 2) DITING があれば突合
    diting_df = load_diting_csv(args.diting)
    if diting_df is not None:
        for p in projects:
            vd_triplets = extract_vd_triplets(p)  # (file,line,sink?)
            compute_matches(p, diting_df, vd_triplets, line_tol=args.diting_line_tol)

    # 3) DataFrame 作成（日本語ヘッダ）
    df_overview = build_overview_df(projects)
    df_projects = build_per_project_df(projects)
    df_token = build_token_time_df(projects)
    df_diting = build_diting_compare_df(projects)
    df_detail = build_vuln_detail_df(projects)
    df_phase  = build_phase_breakdown_df(projects)

    # 4) Excel 出力（日本語シート名 & 列幅自動調整）
    ensure_dir(args.out)
    with pd.ExcelWriter(args.out, engine="openpyxl") as writer:
        write_df("概要", df_overview, writer, freeze=None, apply_autofilter=False)
        write_df("プロジェクト別", df_projects, writer)
        write_df("トークン・時間", df_token, writer)
        write_df("DITING比較", df_diting, writer)
        write_df("検出詳細", df_detail, writer)
        if not df_phase.empty:
            write_df("フェーズ内訳(time.txt)", df_phase, writer)

    print(f"[OK] Excel を出力しました: {args.out}")

if __name__ == "__main__":
    main()
