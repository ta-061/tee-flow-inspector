#!/usr/bin/env python3
"""
TA脆弱性分析結果をExcelにまとめるスクリプト
"""

import json
import glob
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def load_json_files(directory):
    """指定ディレクトリ内の全JSONファイルを読み込む"""
    json_files = glob.glob(os.path.join(directory, "*_vulnerabilities.json"))
    results = []

    for filepath in sorted(json_files):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                ta_name = os.path.basename(filepath).replace('_vulnerabilities.json', '')
                results.append({
                    'ta_name': ta_name,
                    'filepath': filepath,
                    'data': data
                })
        except Exception as e:
            print(f"Error loading {filepath}: {e}")

    return results


def create_summary_sheet(wb, results):
    """サマリーシートを作成"""
    ws = wb.active
    ws.title = "Summary"

    # ヘッダー設定
    headers = [
        "TA Name",
        "Analysis Date",
        "Execution Time (sec)",
        "Time Formatted",
        "Total Flows Analyzed",
        "Flows with Vulnerabilities",
        "Vulnerability Count",
        "Structural Risk Count",
        "Total Findings",
        "LLM Calls",
        "Cache Hit Rate"
    ]

    # ヘッダー行の書き込み
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行の書き込み
    for row, result in enumerate(results, 2):
        data = result['data']
        stats = data.get('statistics', {})

        ws.cell(row=row, column=1, value=result['ta_name'])
        ws.cell(row=row, column=2, value=data.get('analysis_date', 'N/A'))
        ws.cell(row=row, column=3, value=data.get('analysis_time_seconds', 0))
        ws.cell(row=row, column=4, value=data.get('analysis_time_formatted', 'N/A'))
        ws.cell(row=row, column=5, value=stats.get('total_flows_analyzed', 0))
        ws.cell(row=row, column=6, value=stats.get('flows_with_vulnerabilities', 0))
        ws.cell(row=row, column=7, value=data.get('total_vulnerability_lines', 0))
        ws.cell(row=row, column=8, value=data.get('total_finding_lines', 0) - data.get('total_vulnerability_lines', 0))
        ws.cell(row=row, column=9, value=data.get('total_finding_lines', 0))
        ws.cell(row=row, column=10, value=stats.get('llm_calls', 0))
        ws.cell(row=row, column=11, value=stats.get('cache_hit_rate', 'N/A'))

    # 列幅の自動調整
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def create_vulnerabilities_sheet(wb, results):
    """脆弱性詳細シートを作成"""
    ws = wb.create_sheet(title="Vulnerabilities")

    # ヘッダー設定
    headers = [
        "TA Name",
        "Finding ID",
        "Type",
        "File Path",
        "Line Number",
        "Rule ID",
        "Functions",
        "Description",
        "Code Excerpt"
    ]

    # ヘッダー行の書き込み
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行の書き込み
    row = 2
    for result in results:
        ta_name = result['ta_name']
        data = result['data']

        # 脆弱性（Vulnerabilities）を追加
        for vuln in data.get('vulnerabilities', []):
            ws.cell(row=row, column=1, value=ta_name)
            ws.cell(row=row, column=2, value=vuln.get('finding_id', 'N/A'))
            ws.cell(row=row, column=3, value="Vulnerability")
            ws.cell(row=row, column=4, value=vuln.get('file', 'N/A'))
            ws.cell(row=row, column=5, value=vuln.get('line', 'N/A'))
            ws.cell(row=row, column=6, value=vuln.get('primary_rule', 'N/A'))
            ws.cell(row=row, column=7, value=', '.join(vuln.get('functions', [])))

            # 説明を結合
            descriptions = vuln.get('descriptions', [])
            ws.cell(row=row, column=8, value='\n'.join(descriptions) if descriptions else 'N/A')

            # コード抜粋を結合
            code_excerpts = vuln.get('code_excerpts', [])
            ws.cell(row=row, column=9, value='\n'.join(code_excerpts) if code_excerpts else 'N/A')

            row += 1

        # 構造的リスク（Structural Risks）を追加
        for risk in data.get('structural_risks', []):
            ws.cell(row=row, column=1, value=ta_name)
            ws.cell(row=row, column=2, value=risk.get('finding_id', 'N/A'))
            ws.cell(row=row, column=3, value="Structural Risk")
            ws.cell(row=row, column=4, value=risk.get('file', 'N/A'))
            ws.cell(row=row, column=5, value=risk.get('line', 'N/A'))
            ws.cell(row=row, column=6, value=risk.get('primary_rule', 'N/A'))
            ws.cell(row=row, column=7, value=', '.join(risk.get('functions', [])))

            # 説明を結合
            descriptions = risk.get('descriptions', [])
            ws.cell(row=row, column=8, value='\n'.join(descriptions) if descriptions else 'N/A')

            # コード抜粋を結合
            code_excerpts = risk.get('code_excerpts', [])
            ws.cell(row=row, column=9, value='\n'.join(code_excerpts) if code_excerpts else 'N/A')

            row += 1

    # 列幅の調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 60
    ws.column_dimensions['I'].width = 60


def create_statistics_sheet(wb, results):
    """統計情報シートを作成"""
    ws = wb.create_sheet(title="Statistics")

    # ヘッダー設定
    headers = [
        "TA Name",
        "Critical",
        "High",
        "Medium",
        "Low",
        "Total Detections (Before)",
        "Total Lines (After)",
        "Consolidation Rate",
        "Retries"
    ]

    # ヘッダー行の書き込み
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="28B463", end_color="28B463", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行の書き込み
    for row, result in enumerate(results, 2):
        ta_name = result['ta_name']
        stats = result['data'].get('statistics', {})
        severity = stats.get('severity_distribution', {})

        ws.cell(row=row, column=1, value=ta_name)
        ws.cell(row=row, column=2, value=severity.get('critical', 0))
        ws.cell(row=row, column=3, value=severity.get('high', 0))
        ws.cell(row=row, column=4, value=severity.get('medium', 0))
        ws.cell(row=row, column=5, value=severity.get('low', 0))
        ws.cell(row=row, column=6, value=stats.get('total_detections_before_consolidation', 0))
        ws.cell(row=row, column=7, value=stats.get('total_lines_after_consolidation', 0))
        ws.cell(row=row, column=8, value=stats.get('consolidation_rate', 'N/A'))
        ws.cell(row=row, column=9, value=stats.get('retries', 0))

    # 列幅の調整
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def main():
    """メイン処理"""
    script_dir = os.path.dirname(os.path.abspath(__file__))

    print("Loading JSON files...")
    results = load_json_files(script_dir)

    if not results:
        print("No JSON files found!")
        return

    print(f"Found {len(results)} JSON files")

    # Excelワークブックの作成
    wb = Workbook()

    print("Creating Summary sheet...")
    create_summary_sheet(wb, results)

    print("Creating Vulnerabilities sheet...")
    create_vulnerabilities_sheet(wb, results)

    print("Creating Statistics sheet...")
    create_statistics_sheet(wb, results)

    # ファイル保存
    output_file = os.path.join(script_dir, "ta_vulnerabilities_summary.xlsx")
    wb.save(output_file)

    print(f"\nExcel file created: {output_file}")
    print(f"Total TAs analyzed: {len(results)}")

    # 簡易サマリー表示
    total_vulns = sum(r['data'].get('total_vulnerability_lines', 0) for r in results)
    total_risks = sum(r['data'].get('total_finding_lines', 0) - r['data'].get('total_vulnerability_lines', 0) for r in results)

    print(f"Total Vulnerabilities: {total_vulns}")
    print(f"Total Structural Risks: {total_risks}")


if __name__ == "__main__":
    main()
