#!/usr/bin/env python3
"""
TA脆弱性分析結果をExcelにまとめるスクリプト
"""

import json
import glob
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def load_json_files(directory):
    """指定ディレクトリ内の全JSONファイルを読み込む"""
    json_files = glob.glob(os.path.join(directory, "*_vulnerabilities.json"))
    results = []

    for filepath in sorted(json_files):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                ta_name = os.path.basename(filepath).replace('_vulnerabilities.json', '')
                results.append({
                    'ta_name': ta_name,
                    'filepath': filepath,
                    'data': data
                })
        except Exception as e:
            print(f"Error loading {filepath}: {e}")

    return results


def count_unique_findings(data):
    """同一(file, line)で統合した後のユニークな検出数を計算"""
    unique_locations = set()

    # vulnerabilitiesから(file, line)を収集
    for vuln in data.get('vulnerabilities', []):
        file_path = normalize_file_path(vuln.get('file', 'N/A'))
        line = vuln.get('line', 'N/A')
        unique_locations.add((file_path, line))

    # structural_risksから(file, line)を収集
    for risk in data.get('structural_risks', []):
        file_path = normalize_file_path(risk.get('file', 'N/A'))
        line = risk.get('line', 'N/A')
        unique_locations.add((file_path, line))

    return len(unique_locations)


def create_summary_sheet(wb, results):
    """サマリーシートを作成"""
    ws = wb.active
    ws.title = "Summary"

    # ヘッダー設定
    headers = [
        "TA Name",
        "Analysis Date",
        "Execution Time (sec)",
        "Time Formatted",
        "Total Flows Analyzed",
        "Flows with Vulnerabilities",
        "Vulnerability Lines",
        "Structural Risk Lines",
        "Unique Findings (Merged)",
        "LLM Calls",
        "Cache Hit Rate"
    ]

    # ヘッダー行の書き込み
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行の書き込み
    for row, result in enumerate(results, 2):
        data = result['data']
        stats = data.get('statistics', {})

        # 統合後のユニークな検出数を計算
        unique_count = count_unique_findings(data)

        ws.cell(row=row, column=1, value=result['ta_name'])
        ws.cell(row=row, column=2, value=data.get('analysis_date', 'N/A'))
        ws.cell(row=row, column=3, value=data.get('analysis_time_seconds', 0))
        ws.cell(row=row, column=4, value=data.get('analysis_time_formatted', 'N/A'))
        ws.cell(row=row, column=5, value=stats.get('total_flows_analyzed', 0))
        ws.cell(row=row, column=6, value=stats.get('flows_with_vulnerabilities', 0))
        ws.cell(row=row, column=7, value=data.get('total_vulnerability_lines', 0))
        ws.cell(row=row, column=8, value=data.get('total_structural_risk_lines', 0))
        ws.cell(row=row, column=9, value=unique_count)
        ws.cell(row=row, column=10, value=stats.get('llm_calls', 0))
        ws.cell(row=row, column=11, value=stats.get('cache_hit_rate', 'N/A'))

    # 列幅の自動調整
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def normalize_file_path(file_path):
    """ファイルパスを正規化（file:プレフィックスを除去）"""
    if file_path.startswith("file:"):
        return file_path[5:]
    return file_path


def create_vulnerabilities_sheet(wb, results):
    """脆弱性詳細シートを作成（vulnerabilityとstructural_riskを同一行で統合）"""
    ws = wb.create_sheet(title="Vulnerabilities")

    # ヘッダー設定
    headers = [
        "TA Name",
        "Finding ID",
        "Type",
        "File Path",
        "Line Number",
        "Rule ID",
        "Functions",
        "Vulnerability Description",
        "Structural Risk Description",
        "Code Excerpt"
    ]

    # ヘッダー行の書き込み
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データを(TA, file, line)ごとに統合
    row = 2
    for result in results:
        ta_name = result['ta_name']
        data = result['data']

        # (file, line)をキーとした辞書で統合
        merged_findings = {}

        # 脆弱性を追加
        for vuln in data.get('vulnerabilities', []):
            file_path = normalize_file_path(vuln.get('file', 'N/A'))
            line = vuln.get('line', 'N/A')
            key = (file_path, line)

            if key not in merged_findings:
                merged_findings[key] = {
                    'vuln_ids': [],
                    'risk_ids': [],
                    'file': file_path,
                    'line': line,
                    'vuln_rule_ids': set(),
                    'risk_rule_ids': set(),
                    'functions': set(),
                    'vuln_descriptions': [],
                    'risk_descriptions': [],
                    'code_excerpts': set()
                }

            entry = merged_findings[key]
            entry['vuln_ids'].append(vuln.get('vulnerability_id', 'N/A'))

            # Rule IDsを収集（vulnerability側）
            for rule_id in vuln.get('rule_ids', []):
                if rule_id and rule_id != 'N/A':
                    entry['vuln_rule_ids'].add(rule_id)

            # Functionsを収集
            for func in vuln.get('functions', []):
                if func:
                    entry['functions'].add(func)

            # Descriptionsを収集
            for desc in vuln.get('descriptions', []):
                if desc:
                    entry['vuln_descriptions'].append(desc)

            # Code excerptsを収集
            for code in vuln.get('chains', []):
                if code:
                    entry['code_excerpts'].add(' -> '.join(code))

        # 構造的リスクを追加
        for risk in data.get('structural_risks', []):
            file_path = normalize_file_path(risk.get('file', 'N/A'))
            line = risk.get('line', 'N/A')
            key = (file_path, line)

            if key not in merged_findings:
                merged_findings[key] = {
                    'vuln_ids': [],
                    'risk_ids': [],
                    'file': file_path,
                    'line': line,
                    'vuln_rule_ids': set(),
                    'risk_rule_ids': set(),
                    'functions': set(),
                    'vuln_descriptions': [],
                    'risk_descriptions': [],
                    'code_excerpts': set()
                }

            entry = merged_findings[key]
            entry['risk_ids'].append(risk.get('finding_id', 'N/A'))

            # Rule IDsを収集（structural_risk側）
            for rule_id in risk.get('rules', []):
                if rule_id and rule_id != 'N/A' and rule_id != 'other':
                    entry['risk_rule_ids'].add(rule_id)

            # Functionsを収集
            for func in risk.get('functions', []):
                if func:
                    entry['functions'].add(func)

            # Descriptionsを収集
            for desc in risk.get('descriptions', []):
                if desc:
                    entry['risk_descriptions'].append(desc)

            # Code excerptsを収集
            for code in risk.get('code_excerpts', []):
                if code:
                    entry['code_excerpts'].add(code)

        # 統合されたデータをExcelに書き込み
        for (file_path, line), entry in sorted(merged_findings.items(), key=lambda x: (x[0][0], x[0][1])):
            # Finding ID（vulnerability_id / finding_id）
            all_ids = entry['vuln_ids'] + entry['risk_ids']
            finding_id = ', '.join(all_ids) if all_ids else 'N/A'

            # Type（Vulnerability / Structural Risk / Both）
            has_vuln = bool(entry['vuln_ids'])
            has_risk = bool(entry['risk_ids'])
            if has_vuln and has_risk:
                finding_type = "Vulnerability/Structural Risk"
            elif has_vuln:
                finding_type = "Vulnerability"
            else:
                finding_type = "Structural Risk"

            # Rule ID（N/A以外を重複なしで列挙）
            all_rule_ids = sorted(entry['vuln_rule_ids'] | entry['risk_rule_ids'])
            rule_id_str = ', '.join(all_rule_ids) if all_rule_ids else ''

            # Functions（重複なしで列挙）
            functions_str = ', '.join(sorted(entry['functions'])) if entry['functions'] else 'N/A'

            # Vulnerability Description
            vuln_desc = '\n'.join(entry['vuln_descriptions']) if entry['vuln_descriptions'] else ''

            # Structural Risk Description
            risk_desc = '\n'.join(entry['risk_descriptions']) if entry['risk_descriptions'] else ''

            # Code Excerpt
            code_excerpt = '\n'.join(sorted(entry['code_excerpts'])) if entry['code_excerpts'] else 'N/A'

            ws.cell(row=row, column=1, value=ta_name)
            ws.cell(row=row, column=2, value=finding_id)
            ws.cell(row=row, column=3, value=finding_type)
            ws.cell(row=row, column=4, value=file_path)
            ws.cell(row=row, column=5, value=line)
            ws.cell(row=row, column=6, value=rule_id_str)
            ws.cell(row=row, column=7, value=functions_str)
            ws.cell(row=row, column=8, value=vuln_desc)
            ws.cell(row=row, column=9, value=risk_desc)
            ws.cell(row=row, column=10, value=code_excerpt)

            row += 1

    # 列幅の調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 60
    ws.column_dimensions['I'].width = 60
    ws.column_dimensions['J'].width = 60


def create_statistics_sheet(wb, results):
    """統計情報シートを作成"""
    ws = wb.create_sheet(title="Statistics")

    # ヘッダー設定
    headers = [
        "TA Name",
        "Critical",
        "High",
        "Medium",
        "Low",
        "Total Detections (Before)",
        "Total Lines (After)",
        "Consolidation Rate",
        "Retries"
    ]

    # ヘッダー行の書き込み
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="28B463", end_color="28B463", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行の書き込み
    for row, result in enumerate(results, 2):
        ta_name = result['ta_name']
        stats = result['data'].get('statistics', {})
        severity = stats.get('severity_distribution', {})

        ws.cell(row=row, column=1, value=ta_name)
        ws.cell(row=row, column=2, value=severity.get('critical', 0))
        ws.cell(row=row, column=3, value=severity.get('high', 0))
        ws.cell(row=row, column=4, value=severity.get('medium', 0))
        ws.cell(row=row, column=5, value=severity.get('low', 0))
        ws.cell(row=row, column=6, value=stats.get('total_detections_before_consolidation', 0))
        ws.cell(row=row, column=7, value=stats.get('total_lines_after_consolidation', 0))
        ws.cell(row=row, column=8, value=stats.get('consolidation_rate', 'N/A'))
        ws.cell(row=row, column=9, value=stats.get('retries', 0))

    # 列幅の調整
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def main():
    """メイン処理"""
    script_dir = os.path.dirname(os.path.abspath(__file__))

    print("Loading JSON files...")
    results = load_json_files(script_dir)

    if not results:
        print("No JSON files found!")
        return

    print(f"Found {len(results)} JSON files")

    # Excelワークブックの作成
    wb = Workbook()

    print("Creating Summary sheet...")
    create_summary_sheet(wb, results)

    print("Creating Vulnerabilities sheet...")
    create_vulnerabilities_sheet(wb, results)

    print("Creating Statistics sheet...")
    create_statistics_sheet(wb, results)

    # ファイル保存
    output_file = os.path.join(script_dir, "ta_vulnerabilities_summary.xlsx")
    wb.save(output_file)

    print(f"\nExcel file created: {output_file}")
    print(f"Total TAs analyzed: {len(results)}")

    # 簡易サマリー表示
    total_vulns = sum(r['data'].get('total_vulnerability_lines', 0) for r in results)
    total_risks = sum(r['data'].get('total_structural_risk_lines', 0) for r in results)
    total_unique = sum(count_unique_findings(r['data']) for r in results)

    print(f"Total Vulnerability Lines: {total_vulns}")
    print(f"Total Structural Risk Lines: {total_risks}")
    print(f"Total Unique Findings (Merged): {total_unique}")


if __name__ == "__main__":
    main()
